import csv
import os
import openpyxl

Base_Repository = 'Bigdata_Repository'
Type_A_Repository = 'TypeA'
Type_A_format = 'csv'
Type_A_upper_limit_of_bulk = 10000
Type_B_Repository = 'TypeB'
Type_B_format = 'xlsx'
Type_B_upper_limit_of_bulk = 20000

demo_data = ['1111', '동구', '대구광역시', '아양철교', '5', '35252', '55104']
simulation_count = 100

Type_A_file_name = 'csv_demo_file_'
Type_B_file_name = 'excel_demo_file_'


def data_settings():

    global Base_Repository, Type_A_Repository, Type_A_format, Type_A_upper_limit_of_bulk, Type_B_Repository, Type_B_format, Type_B_upper_limit_of_bulk
    while True:
        print("\t\t\t\t환경 설정")
        print("=======================================")
        print(f"1. Base_Repository: {Base_Repository}")
        print(f"2. Type_A Repository: {Type_A_Repository}")
        print(f"   Type_A format: {Type_A_format}")
        print("3. Type_A 데이터 용량제한:", Type_A_upper_limit_of_bulk)
        print(f"4. Type_B Repository: {Type_B_Repository}")
        print(f"   Type_B format: {Type_B_format}")
        print("5. Type_B 데이터 용량제한:", Type_B_upper_limit_of_bulk)
        print("6. 이전메뉴")
        setting_option = input(">>> ")
        if setting_option == '1':
            Base_Repository = input("Base_Repository 변경할 이름을 입력하세요: ")
        elif setting_option == '2':
            Type_A_Repository = input("Type A 저장소 이름을 설정하세요: ")
            Type_A_format = input("Type A 형식을 입력하세요: ")
        elif setting_option == '3':
            Type_A_upper_limit_of_bulk = input("Type_A 데이터 용량제한을 설정하세요: ")
        elif setting_option == '4':
            Type_B_Repository = input("Type B 저장소 이름을 설정하세요: ")
            Type_B_format = input("Type B 형식을 입력하세요: ")
        elif setting_option == '5':
            Type_A_upper_limit_of_bulk = input("Type_B 데이터 용량제한을 설정하세요: ")
        elif setting_option == '6':
            return
        else:
            print("올바른 명령이 아닙니다.")


def data_save_simulation():

    if not os.path.exists(Base_Repository):
        os.mkdir(Base_Repository)
        print("해당 디렉토리에 Base_Repository가 없어 새로만들었습니다.")

    save_option = '''   데이터 수집 옵션
    ====================
    1. Type A 데이터 수집
    2. Type B 데이터 수집
    3. 이전 메뉴
    ====================
    >>> '''
    while True:
        save_option_selection = input(save_option)
        if save_option_selection == '1':
            save_csv_format()
        elif save_option_selection == '2':
            save_excel_format()
        elif save_option_selection == '3':
            return
        else:
            print('올바른 명령이 아닙니다.')


def save_csv_format():

    is_first = False
    is_full = False
    is_header = False
    header_list = ['addrCd', 'gungu', 'city', 'resNm', 'rnum', 'csForCnt', 'csNatCnt']

    if not os.path.exists(f"{Base_Repository}/{Type_A_Repository}"):
        os.mkdir(f"{Base_Repository}/{Type_A_Repository}")
        print("해당 디렉토리에 Type_A_Repository가 없어 새로만들었습니다.")

    csv_dir = f"{Base_Repository}/{Type_A_Repository}"

    csv_index = len(os.listdir(csv_dir))
    if csv_index == 0:
        is_first = True
    else:
        last_file_size = os.path.getsize(csv_dir + f"/{Type_A_file_name}{csv_index}.{Type_A_format}")
        print(csv_dir + f"/{Type_A_file_name}{csv_index}.{Type_A_format} 파일 크기: ", last_file_size)
        print('파일 크기 상한:', Type_A_upper_limit_of_bulk)
        if last_file_size > Type_A_upper_limit_of_bulk:
            print('파일 크기가 임계점을 초과하여 새로운 파일을 생성합니다.')
            is_full = True

    if is_first or is_full:
        csv_index += 1
        is_header = True

    file_name = csv_dir + f"/{Type_A_file_name}{csv_index}.{Type_A_format}"
    csv_out_file = open(file_name, 'a', newline='')
    file_writer = csv.writer(csv_out_file)
    if is_header:
        file_writer.writerow(header_list)

    for index in range(simulation_count):
        file_writer.writerow(demo_data)
    csv_out_file.close()


def save_excel_format():

    is_first = False
    is_full = False
    header_list = ['addrCd', 'gungu', 'city', 'resNm', 'rnum', 'csForCnt', 'csNatCnt']

    if not os.path.exists(f"{Base_Repository}/{Type_B_Repository}"):
        os.mkdir(f"{Base_Repository}/{Type_B_Repository}")
        print("해당 디렉토리에 Type_B_Repository가 없어 새로만들었습니다.")

    excel_dir = f"{Base_Repository}/{Type_B_Repository}"

    excel_index = len(os.listdir(excel_dir))
    if excel_index == 0:
        is_first = True
    else:
        last_file_size = os.path.getsize(excel_dir + f"/{Type_B_file_name}{excel_index}.{Type_B_format}")
        print(excel_dir + f"/{Type_B_file_name}{excel_index}.{Type_A_format} 파일 크기: ", last_file_size)
        print('파일 크기 상한:', Type_B_upper_limit_of_bulk)
        if last_file_size > Type_B_upper_limit_of_bulk:
            print('파일 크기가 임계점을 초과하여 새로운 파일을 생성합니다.')
            is_full = True

    if is_first or is_full:
        new_workbook = openpyxl.Workbook()
        sheet1 = new_workbook['Sheet']
        sheet1.title = 'Excel_Demo'
        sheet1.append(header_list)

        for index in range(simulation_count):
            sheet1.append(demo_data)
        new_workbook.save(excel_dir + f"/{Type_B_file_name}{excel_index + 1}.{Type_B_format}")

    else:
        load_workbook = openpyxl.load_workbook(excel_dir + f"/{Type_B_file_name}{excel_index}.{Type_B_format}")
        sheet1_loaded = load_workbook.active
        for index in range(simulation_count):
            sheet1_loaded.append(demo_data)

        load_workbook.save(excel_dir + f"/{Type_B_file_name}{excel_index}.{Type_B_format}")


def main():
    option_first = '''>>>>>저장소 시뮬레이션<<<<<
  ======================
\t1. 환경  설정
\t2. 작업  수행   
\t3. 종료     
  ======================
>>> '''
    while True:
        selected_first_option = input(option_first)

        if selected_first_option == '1':
            data_settings()
        elif selected_first_option == '2':
            data_save_simulation()
        elif selected_first_option == '3':
            print("프로그램이 종료됩니다.")
            return
        else:
            print("올바른 명령이 아닙니다.")


if __name__ == '__main__':
    main()
